import openpyxl

while True:
    # Kullanıcıdan makine bilgisini alma
    machine = input("\nLütfen Makine Seçiniz (Plazma için 'P', Bystronic için 'B'): ").upper()

    # Makine bilgisinin doğruluğunu kontrol etme
    if machine not in ['P', 'B']:
        print("Hatalı makine bilgisi! Lütfen 'P' veya 'B' giriniz.")
        continue  # Döngünün başına dön

    # Excel dosyasını açma
    wb = openpyxl.load_workbook("data.xlsx")

    # Makineye göre sayfayı seçme
    if machine == "P":
        sheet = wb["MESSER"]
        machine_name = "Plazma (MESSER)"
    else:
        sheet = wb["BYSTRONIC"]
        machine_name = "Bystronic"

    # Kalınlık bilgilerini ekrana yazdırma
    print(f"\n{machine_name} Makinesine Ait Veritabanındaki Kalınlık Bilgileri:")
    thickness_list = [str(sheet.cell(row=i, column=1).value) + "mm," for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]  # Kalınlık sütunundaki boş hücreleri atlayarak alıyoruz
    print(" ".join(thickness_list))

    # Kullanıcıdan kalınlık bilgisini alma
    thickness_input = input("\nLütfen Kalınlığı Giriniz: ")

    # Kalınlık bilgisine denk gelen satırı bulma
    found = False
    for row in range(2, sheet.max_row + 1):
        thickness_cell = sheet.cell(row=row, column=1).value
        if thickness_cell == thickness_input:  # Kalınlık sütunu 1. sütun olduğu varsayılarak alınıyor
            found = True
            thickness = str(thickness_cell) + " mm"
            break

    # Sonuçları yazdırma
    if found:
        speed_column_index = None
        hole_speed_column_index = None
        piercing_time_column_index = None
        machine_speed_column_index = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == "KONTUR_KESIM_HIZI":
                speed_column_index = col
            elif sheet.cell(row=1, column=col).value == "DELIK_KESIM_HIZI":
                hole_speed_column_index = col
            elif sheet.cell(row=1, column=col).value == "PIERCING_TIME":
                piercing_time_column_index = col
            elif sheet.cell(row=1, column=col).value == "MAKINE_HIZI":
                machine_speed_column_index = col

        if speed_column_index is None or hole_speed_column_index is None or piercing_time_column_index is None or machine_speed_column_index is None:
            print("Hız, Delik Kesim Hızı, Piercing Time veya Makine Hızı sütunu bulunamadı!")
            exit()

        speed = sheet.cell(row=row, column=speed_column_index).value
        hole_speed = sheet.cell(row=row, column=hole_speed_column_index).value
        piercing_time = sheet.cell(row=row, column=piercing_time_column_index).value
        machine_speed = sheet.cell(row=row, column=machine_speed_column_index).value
        print(f"\n{thickness} Kalınlık İçin:\n"
              f"Kesim Hızı: {speed} mm/dk\n"
              f"Delik Kesim Hızı: {hole_speed} mm/dk\n"
              f"Piercing Time: {piercing_time} sn\n"
              f"Makine Hızı: {machine_speed} mm/dk\n")
        break  # Doğru sonuç bulundu, döngüden çık
    else:
        print(f"Kalınlık değeri ({thickness_input}) bulunamadı!")
